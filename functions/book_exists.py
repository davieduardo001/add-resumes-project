import openpyxl
import os

####################TESTS
#verify if the book exists
def does_file_exists(file_neme):
    return os.path.exists(file_neme)

if does_file_exists('./planilha.xlsx'):
    print('open the book')
    book=openpyxl.load_workbook('planilha.xlsx')
else:
    book=openpyxl.Workbook()
    book.create_sheet('MAIN')
    page=book['MAIN']
    page.append(['NOME','SEXO','LINK-GOOGLE','LINK-LATTES'])
    book.save('planilha.xlsx')
    os.system('cls')