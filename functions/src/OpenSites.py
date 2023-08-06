import openpyxl
import os

####WITHOUT-URL-FUNCTIONS####
##############LATTES
def lattes(p, name):
    browser=p.chromium.launch(headless=False)                           #open the browser
    context=browser.new_context()                                       #create a new context

    page=context.new_page()

    page.goto("http://buscatextual.cnpq.br/buscatextual/busca.do?metodo=apresentar")

    page.locator('xpath=//*[@id="textoBusca"]').fill(name)              #fill the name#
    page.locator('xpath=//*[@id="buscarDemais"]').click()               #select "more options"
    page.locator('a.button#botaoBuscaFiltros').click()                  #search on lattes

    print('\nwaiting... ')                                          
    response=input('Did you find the resume? [y/N] ')                         
   
    if response=='y' or response=='Y':
        with context.expect_page() as resume_page_info:                     #open resume###
            page.locator('xpath=//*[@id="idbtnabrircurriculo"]').click()    ###############
        resume_page = resume_page_info.value                                ###############
        resume_page.wait_for_load_state()                                   ###############

        global lattes_url                                                   #set the var###
        lattes_url=resume_page.url                                          #take the url##
        page.reload()

        #SEARCH ON LATTES
        #uniser
        label_uniser=resume_page.locator('text=UniSER')
        count_uniser=label_uniser.count()
        #unb
        label_unb=resume_page.locator('text=UnB')
        count_unb=label_unb.count()
        #elder
        label_elder=resume_page.locator('text=elder')
        count_elder=label_elder.count()
        #envelhecimento
        label_envelhecimento=resume_page.locator('text=envelhecimento')
        count_envelhecimento=label_envelhecimento.count()
        #envelhecer
        label_envelhecer=resume_page.locator('text=envelhecer')
        count_envelhecer=label_envelhecer.count()
        #aging
        label_aging=resume_page.locator('text=aging')
        count_aging=label_aging.count()
        #older
        label_older=page.locator('text=older')
        count_older=label_older.count()
        #senescence
        label_senescence=page.locator('text=senescence')
        count_senescence=label_senescence.count()
        #senescência
        label_senescencia=page.locator('text=senescência')
        count_senescencia=label_senescencia.count()
        #seniors
        label_seniors=page.locator('text=seniors')
        count_seniors=label_seniors.count()

        ##SAVING##
        total_of_articles=count_uniser+count_unb+count_elder+count_aging+count_envelhecimento+count_envelhecer+count_senescence+count_senescencia+count_seniors+count_older

        print('\nthe total of articles are: ', total_of_articles)
        input('saving the LATTES profile, please ENTER ')
        #CREATING USER SHEET
        book=openpyxl.load_workbook('planilha.xlsx')
        book.create_sheet(name)
        book.save('planilha.xlsx')
        user_page=book[name]
        user_page.append([name])
        user_page.append(['####'])
        user_page.append(['Resumo:'])
        user_page.append([resume_page.locator('xpath=/html/body/div[1]/div[3]/div/div/div/div[2]/p').text_content()])
        user_page.append(['####'])
        user_page.append(['Lattes curriculo: '])

        #SENDING ARTICLES
        for i in range(count_uniser):                                   #uniser
            user_page.append([label_uniser.nth(i).text_content()])
        for i in range(count_unb):                                      #unb
            user_page.append([label_unb.nth(i).text_content()])
        for i in range(count_elder):                                    #elder
            user_page.append([label_elder.nth(i).text_content()])  
        for i in range(count_aging):                                    #aging
            user_page.append([label_aging.nth(i).text_content()])    
        for i in range(count_envelhecimento):                           #envelhecimento
            user_page.append([label_envelhecimento.nth(i).text_content()])
        for i in range(count_older):                                    #older
            user_page.append([label_older.nth(i).text_content()])
        for i in range(count_envelhecer):                               #envelhecer            
            user_page.append([label_envelhecer.nth(i).text_content()])
        for i in range(count_seniors):                                   #seniors            
            user_page.append([label_seniors.nth(i).text_content()]) 
        for i in range(count_senescencia):                              #senescência            
            user_page.append([label_senescencia.nth(i).text_content()])
        for i in range(count_senescence):                               #senescence            
            user_page.append([label_senescence.nth(i).text_content()]) 
        book.save('planilha.xlsx')
        browser.close()

    else:
        print('okay, this person dont have the Lattes resume...')
        lattes_url='---'
        browser.close()
        book=openpyxl.load_workbook('planilha.xlsx')
        book.create_sheet(name)
        book.save('planilha.xlsx') 

##############GOOGLE
def google(p, name):
    print('\nLoading GOOGLE resume...')
    browser=p.chromium.launch(headless=False)                           #open the browser
    context=browser.new_context()                                       #create a new context

    page=context.new_page()

    page.goto("https://scholar.google.com.br/citations?hl=pt-BR&view_op=search_authors&mauthors=&btnG=")

    page.locator('xpath=//*[@id="gs_hdr_tsi"]').fill(name)              #fill the name#
    page.locator('xpath=//*[@id="gs_hdr_tsb"]').click()                 #search########

    print('\nwaiting... ')
    response=input('Did you find the resume? [y/N] ')

    if response=='Y' or response=='y':
        input('MAKE SURE that all articles are showing. ')
        page.reload()

        #SEARCH ON GOOGLE
        #uniser
        label_uniser=page.locator('text=UniSER')
        count_uniser=label_uniser.count()
        #unb
        label_unb=page.locator('text=UnB')
        count_unb=label_unb.count()
        #elder
        label_elder=page.locator('text=elder')
        count_elder=label_elder.count()
        #envelhecimento
        label_envelhecimento=page.locator('text=envelhecimento')
        count_envelhecimento=label_envelhecimento.count()
        #envelhecer
        label_envelhecer=page.locator('text=envelhecer')
        count_envelhecer=label_envelhecer.count()
        #aging
        label_aging=page.locator('text=aging')
        count_aging=label_aging.count()
        #older
        label_older=page.locator('text=older')
        count_older=label_older.count()
        #senescence
        label_senescence=page.locator('text=senescence')
        count_senescence=label_senescence.count()
        #senescência
        label_senescencia=page.locator('text=senescência')
        count_senescencia=label_senescencia.count()
        #seniors
        label_seniors=page.locator('text=seniors')
        count_seniors=label_seniors.count()

        ##SAVING##
        total_of_articles=count_uniser+count_unb+count_elder+count_aging+count_envelhecimento+count_envelhecer+count_senescence+count_senescencia+count_seniors+count_older

        print('\nthe total of articles are: ', total_of_articles)
        response=input('saving the GOOGLE profile, please press ENTER ')
        response.lower()
        page.reload()

        global google_url
        google_url=page.url

        #CREATING USER SHEET
        book=openpyxl.load_workbook('planilha.xlsx')
        user_page=book[name]
        user_page.append(['####'])
        user_page.append(['Google curriculo:'])
        user_page.append(['####'])

        #SENDING ARTICLES
        for i in range(count_uniser):                                   #uniser
            user_page.append([label_uniser.nth(i).text_content()])
        for i in range(count_unb):                                      #unb
            user_page.append([label_unb.nth(i).text_content()])
        for i in range(count_elder):                                    #elder
            user_page.append([label_elder.nth(i).text_content()])  
        for i in range(count_aging):                                    #aging
            user_page.append([label_aging.nth(i).text_content()])    
        for i in range(count_envelhecimento):                           #envelhecimento
            user_page.append([label_envelhecimento.nth(i).text_content()])
        for i in range(count_older):                                    #older
            user_page.append([label_older.nth(i).text_content()])
        for i in range(count_envelhecer):                               #envelhecer            
            user_page.append([label_envelhecer.nth(i).text_content()])
        for i in range(count_seniors):                                   #seniors            
            user_page.append([label_seniors.nth(i).text_content()]) 
        for i in range(count_senescencia):                              #senescência            
            user_page.append([label_senescencia.nth(i).text_content()])
        for i in range(count_senescence):                               #senescence            
            user_page.append([label_senescence.nth(i).text_content()]) 
        book.save('planilha.xlsx')
        browser.close()
    else:
        print('okay, this person dont have the GOOGLE resume...')
        google_url='---'
        browser.close()

#SEND TO THE MAIN SHEET
#obs: this functions depends of the URL in the other 2 functions
def send_to_main_sheet(name, gender):
    book=openpyxl.load_workbook('planilha.xlsx')
    main_page=book['MAIN']
    main_page.append([name, gender, google_url, lattes_url])
    book.save('planilha.xlsx')

####WITH-URL-FUNCTIONS####
def lattes_with_url(p, name, lattesUrl):                        

    if lattesUrl!='':
        browser=p.chromium.launch(headless=False)                           #open the browser
        context=browser.new_context()                                       #create a new context

        page=context.new_page()

        page.goto(lattesUrl) 
        print('\nwaiting... ')                                          
        response=input('when you finished the captcha return and press ENTER ')

        global lattes_url                                                   #set the var###
        lattes_url=page.url                                          #take the url##

        #SEARCH ON LATTES
        #uniser
        label_uniser=resume_page.locator('text=UniSER')
        count_uniser=label_uniser.count()
        #unb
        label_unb=resume_page.locator('text=UnB')
        count_unb=label_unb.count()
        #elder
        label_elder=resume_page.locator('text=elder')
        count_elder=label_elder.count()
        #envelhecimento
        label_envelhecimento=resume_page.locator('text=envelhecimento')
        count_envelhecimento=label_envelhecimento.count()
        #envelhecer
        label_envelhecer=resume_page.locator('text=envelhecer')
        count_envelhecer=label_envelhecer.count()
        #aging
        label_aging=resume_page.locator('text=aging')
        count_aging=label_aging.count()
        #older
        label_older=page.locator('text=older')
        count_older=label_older.count()
        #senescence
        label_senescence=page.locator('text=senescence')
        count_senescence=label_senescence.count()
        #senescência
        label_senescencia=page.locator('text=senescência')
        count_senescencia=label_senescencia.count()
        #seniors
        label_seniors=page.locator('text=seniors')
        count_seniors=label_seniors.count()

        ##SAVING##
        total_of_articles=count_uniser+count_unb+count_elder+count_aging+count_envelhecimento+count_envelhecer+count_senescence+count_senescencia+count_seniors+count_older

        print('\nthe total of articles are: ', total_of_articles)
        input('saving the LATTES profile, please ENTER ')
        #CREATING USER SHEET
        book=openpyxl.load_workbook('planilha.xlsx')
        book.create_sheet(name)
        book.save('planilha.xlsx')
        user_page=book[name]
        user_page.append([name])
        user_page.append(['####'])
        user_page.append(['Resumo:'])
        user_page.append([page.locator('xpath=/html/body/div[1]/div[3]/div/div/div/div[2]/p').text_content()])
        user_page.append(['####'])
        user_page.append(['Lattes curriculo: '])

        #SENDING ARTICLES
        for i in range(count_uniser):                                   #uniser
            user_page.append([label_uniser.nth(i).text_content()])
        for i in range(count_unb):                                      #unb
            user_page.append([label_unb.nth(i).text_content()])
        for i in range(count_elder):                                    #elder
            user_page.append([label_elder.nth(i).text_content()])  
        for i in range(count_aging):                                    #aging
            user_page.append([label_aging.nth(i).text_content()])    
        for i in range(count_envelhecimento):                           #envelhecimento
            user_page.append([label_envelhecimento.nth(i).text_content()])
        for i in range(count_older):                                    #older
            user_page.append([label_older.nth(i).text_content()])
        for i in range(count_envelhecer):                               #envelhecer            
            user_page.append([label_envelhecer.nth(i).text_content()])
        for i in range(count_seniors):                                   #seniors            
            user_page.append([label_seniors.nth(i).text_content()]) 
        for i in range(count_senescencia):                              #senescência            
            user_page.append([label_senescencia.nth(i).text_content()])
        for i in range(count_senescence):                               #senescence            
            user_page.append([label_senescence.nth(i).text_content()]) 
        book.save('planilha.xlsx')
        browser.close()
    else:
        print('\nokay, this person dont have the Lattes resume...')
        lattes_url='---'
        book=openpyxl.load_workbook('planilha.xlsx')
        book.create_sheet(name)
        book.save('planilha.xlsx') 

def google_with_url(p, name, googleUrl):

    if googleUrl!='':
        browser=p.chromium.launch(headless=False)                           #open the browser
        context=browser.new_context()                                       #create a new context

        page=context.new_page()

        page.goto(googleUrl)

        print('\nLoading GOOGLE resume...')

        input('MAKE SURE that all articles are showing. ')
        page.reload()

        #SEARCH ON GOOGLE
        #uniser
        label_uniser=page.locator('text=UniSER')
        count_uniser=label_uniser.count()
        #unb
        label_unb=page.locator('text=UnB')
        count_unb=label_unb.count()
        #elder
        label_elder=page.locator('text=elder')
        count_elder=label_elder.count()
        #envelhecimento
        label_envelhecimento=page.locator('text=envelhecimento')
        count_envelhecimento=label_envelhecimento.count()
        #envelhecer
        label_envelhecer=page.locator('text=envelhecer')
        count_envelhecer=label_envelhecer.count()
        #aging
        label_aging=page.locator('text=aging')
        count_aging=label_aging.count()
        #older
        label_older=page.locator('text=older')
        count_older=label_older.count()
        #senescence
        label_senescence=page.locator('text=senescence')
        count_senescence=label_senescence.count()
        #senescência
        label_senescencia=page.locator('text=senescência')
        count_senescencia=label_senescencia.count()
        #seniors
        label_seniors=page.locator('text=seniors')
        count_seniors=label_seniors.count()

        ##SAVING##
        total_of_articles=count_uniser+count_unb+count_elder+count_aging+count_envelhecimento+count_envelhecer+count_senescence+count_senescencia+count_seniors+count_older

        print('\nthe total of articles are: ', total_of_articles)
        response=input('saving the GOOGLE profile, please press ENTER ')
        response.lower()
        page.reload()

        global google_url
        google_url=page.url

        #CREATING USER SHEET
        book=openpyxl.load_workbook('planilha.xlsx')
        user_page=book[name]
        user_page.append(['####']) 
        user_page.append(['Google curriculo:'])
        user_page.append(['####'])

        #SENDING ARTICLES        
        for i in range(count_uniser):                                   #uniser
            user_page.append([label_uniser.nth(i).text_content()])
        for i in range(count_unb):                                      #unb
            user_page.append([label_unb.nth(i).text_content()])
        for i in range(count_elder):                                    #elder
            user_page.append([label_elder.nth(i).text_content()])  
        for i in range(count_aging):                                    #aging
            user_page.append([label_aging.nth(i).text_content()])    
        for i in range(count_envelhecimento):                           #envelhecimento
            user_page.append([label_envelhecimento.nth(i).text_content()])
        for i in range(count_older):                                    #older
            user_page.append([label_older.nth(i).text_content()])
        for i in range(count_envelhecer):                               #envelhecer            
            user_page.append([label_envelhecer.nth(i).text_content()])
        for i in range(count_seniors):                                   #seniors            
            user_page.append([label_seniors.nth(i).text_content()]) 
        for i in range(count_senescencia):                              #senescência            
            user_page.append([label_senescencia.nth(i).text_content()])
        for i in range(count_senescence):                               #senescence            
            user_page.append([label_senescence.nth(i).text_content()]) 
        book.save('planilha.xlsx')
        browser.close()
    else:
        print('\nokay, this person dont have the GOOGLE resume...')
        google_url='---'
