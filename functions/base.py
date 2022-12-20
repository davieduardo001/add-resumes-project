import openpyxl

#create a book
book = openpyxl.Workbook()

#view created pages
print(book.sheetnames)

#create a page
book.create_sheet('MAIN')

#selecting a page
page1 = book['MAIN']
page1.append(['NOME','SEXO','LINK-GOOGLE','LINK-LATTES','LINK-ARTIGOS','ANO-PUBLICACAO'])

#save the book
book.save('testes.xlsx')

#open book
book = openpyxl.load_workbook('testes.xlsx')

#select the page
main_page = book['MAIN']

#printing the data from main
for rows in main_page.iter_rows(min_row=2, max_row=5):
    for cell in rows:
        print(cell.value)